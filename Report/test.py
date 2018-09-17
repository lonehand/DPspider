# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from pandas import DataFrame


# ==========体验报告报表============ #
def CommentUpdate(MidSheet, CommentSheet, CommentSheet_R):
    comnum = 0
    comnum_r = 0
    for i in CommentSheet['B']:
        if i.value == 9:
            comnum += 1
        allcomnum += 1
    for i in CommentSheet_R['B']:
        if i.value == 9:
            comnum_r += 1
    lastcomnum = CommentMaxRow-comnum
    lastcomnum_r = Comment_RMaxRow-comnum_r
    print(comnum, comnum_r, lastcomnum, lastcomnum_r)


def TimeUpdate(MidSheet, TimeInfo):
    row = 2
    num = 0
    Timelist = []
    OldTime = TimeInfo[0]
    NowTime = TimeInfo[1]
    Timelist.append(str(int(NowTime[:4]))+'年')
    Timelist.append(str(int(OldTime[:4]))+'年')
    Timelist.append(str(int(NowTime[5:7]))+'月')
    Timelist.append(str(int(OldTime[5:7]))+'月')
    for col in range(1, 5):
        MidSheet.cell(row, col, Timelist[num])
        num += 1



FileName = 'NewReport/测试.xlsx'
WorkBook = load_workbook(FileName)
MidSheet = WorkBook['MidSheet']
FlowSheet = WorkBook['流量']
ChatSheet = WorkBook['咨询明细']
APPSheet = WorkBook['预约数据']
SaleSheet = WorkBook['消费数据明细（线上）']
CommentSheet = WorkBook['口碑数据']
CommentSheet_R = WorkBook['回复口碑']
# TimeUpdate(MidSheet, TimeInfo)
# CommentUpdate(MidSheet, CommentSheet, CommentSheet_R)
for i in CommentSheet['B']:
	print(i)