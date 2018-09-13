# -*- coding: utf-8 -*-

import datetime
from openpyxl import load_workbook


# 获得最大行数
def GetBooklen(sheetname):
    flowbooklen = sheetname.max_row
    return flowbooklen


# 获得昨天
def Get_yesterday():
    today = datetime.datetime.now()
    onedaydelay = datetime.timedelta(days=1)
    yesterday = (
        today - onedaydelay
        ).strftime(r'%Y-%m-%d')
    return yesterday


# 流量数据管理
def Flowupdate(Data, FlowSheet):
    YesterDay = Get_yesterday()
    MaxLen = GetBooklen(FlowSheet)
    LastDay = FlowSheet.cell(MaxLen, 3).value.strftime('%Y-%m-%d')
    if YesterDay == LastDay:
        pass
    else:
        for data in Data:
            if data > LastDay:
                FlowSheet.append([
                    int(data[:4]),
                    int(data[5:7]),
                    datetime.datetime.strptime(data, "%Y-%m-%d"),
                    int(Data[data][0]),
                    int(Data[data][1]),
                    float('%.2f' % float(Data[data][2])),
                    float('%.2f' % float(Data[data][3])),
                ])


# 口碑数据管理
def ChatUpdate(chatreuslt, ChatSheet):
    row = 1
    for data in chatreuslt:
        row += 1
        num = 0
        for i in range(1, 8):
            ChatSheet.cell(row, i, value=chatreuslt[data][num])
            num += 1


# 订单中心
def AppointUpdate(appointmentresult, APPSheet):
    for data in appointmentresult:
        num = 0
        for col in range(1, 10):
            APPSheet.cell(int(data), col, appointmentresult[data][num])
            num += 1


# 线上销售数据
def SaleOnlineUpdate(SaleOnlineresult, SaleSheet):
    row = 2
    for data in SaleOnlineresult:
        num = 0
        for col in range(1, 14):
            SaleSheet.cell(row, col, SaleOnlineresult[data][num])
            num += 1
        row += 1


# 预约数据
def CommentUpdate(CommentResult, CommentSheet, CommentSheet_R):
    row = 2
    row_r = 2
    for data in CommentResult:
        num = 0
        num_r = 0
        if CommentResult[data][-2] == '是':
            for col in range(1, 16):
                CommentSheet_R.cell(row_r, col, CommentResult[data][num_r])
                num_r += 1
            row_r += 1
        for col in range(1, 16):
            CommentSheet.cell(row, col, CommentResult[data][num])
            num += 1
        row += 1


def Report_main(flowresult, chatresult, appointmentresult, SaleOnlineresult, CommentResult, acount):
    filename = 'Report/%s.xlsx' % acount
    savename = 'Report/NewReport/%s.xlsx' % acount
    WorkBook = load_workbook(filename)
    FlowSheet = WorkBook['流量']
    ChatSheet = WorkBook['咨询明细']
    APPSheet = WorkBook['预约数据']
    SaleSheet = WorkBook['消费数据明细（线上）']
    CommentSheet = WorkBook['口碑数据']
    CommentSheet_R = WorkBook['回复口碑']

    Flowupdate(flowresult, FlowSheet)
    ChatUpdate(chatresult, ChatSheet)
    AppointUpdate(appointmentresult, APPSheet)
    SaleOnlineUpdate(SaleOnlineresult, SaleSheet)
    if CommentResult != 'null':
        CommentUpdate(CommentResult, CommentSheet, CommentSheet_R)
    WorkBook.save(savename)