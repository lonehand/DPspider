# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import calendar


# ==========体验报告报表============ #
def CommentUpdate(MidSheet, CommentSheet, CommentSheet_R, month):
    num = 0
    CommentList = []
    ThisMonthcomment = 0
    LastMonthcomment = 0
    ThisMonthreplay = 0
    LastMonthreplay = 0
    ThisMonthHighScores = 0
    ThisMonthLowScores = 0
    LastMonthHighScores = 0
    LastMonthLowScores = 0
    TotalComment = CommentSheet.max_row + 1
    ToTalReplay = CommentSheet_R.max_row + 1

    for row in range(2, TotalComment):
        if CommentSheet['B%s' % str(row)].value == month:
            ThisMonthcomment += 1
            if CommentSheet['H%s' % str(row)].value > 3:
                ThisMonthHighScores += 1
            else:
                ThisMonthLowScores += 1
        else:
            LastMonthcomment += 1
            if CommentSheet['H%s' % str(row)].value > 3:
                LastMonthHighScores += 1
            else:
                LastMonthLowScores += 1

    for row in range(2, ToTalReplay):
        if CommentSheet_R['B%s' % str(row)].value == month:
            ThisMonthreplay += 1
        else:
            LastMonthreplay += 1

    CommentList.append(ThisMonthcomment)
    CommentList.append(LastMonthcomment)
    CommentList.append(ThisMonthreplay)
    CommentList.append(LastMonthreplay)
    CommentList.append(ThisMonthHighScores)
    CommentList.append(LastMonthHighScores)
    CommentList.append(ThisMonthLowScores)
    CommentList.append(LastMonthLowScores)

    for row in range(5, 9):
        for col in range(2, 4):
            MidSheet.cell(row, col, CommentList[num])
            num += 1


def TimeUpdate(MidSheet, TimeInfo):
    row = 2
    num = 0
    Timelist = []
    OldTime = TimeInfo[0]
    NowTime = TimeInfo[1]
    Timelist.append(str(int(NowTime[:4]))+'年')
    Timelist.append(str(int(OldTime[:4]))+'年')
    Timelist.append(str(int(NowTime[5:7]))+'月')
    Timelist.append(str(int(OldTime[5:7]))+'月')
    Timelist.append(int(NowTime[8:10]))
    Timelist.append(int(calendar.monthrange(int(OldTime[:4]), int(OldTime[5:7]))[1]))
    for col in range(1, 7):
        MidSheet.cell(row, col, Timelist[num])
        num += 1
    return int(NowTime[5:7])


def main(TimeInfo, acount):
    FileName = 'Report/NewReport/%s.xlsx' % acount
    WorkBook = load_workbook(FileName)
    MidSheet = WorkBook['MidSheet']
    FlowSheet = WorkBook['流量']
    ChatSheet = WorkBook['咨询明细']
    APPSheet = WorkBook['预约数据']
    SaleSheet = WorkBook['消费数据明细（线上）']
    CommentSheet = WorkBook['口碑数据']
    CommentSheet_R = WorkBook['回复口碑']
    month = TimeUpdate(MidSheet, TimeInfo)
    CommentUpdate(MidSheet, CommentSheet, CommentSheet_R, month)
    WorkBook.save(FileName)